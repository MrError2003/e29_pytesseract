"""
Módulo de procesamiento: descarga imágenes, aplica OCR multi-transformación
y calcula coincidencias con los datos del Excel con matching aproximado.

Pipeline de imagen (adaptado del enfoque PHP con Intervention Image):
  - 9 variantes de preprocesamiento × 6 modos PSM de Tesseract
  - Corrección automática de rotación (deskew)
  - Parada anticipada cuando se obtiene texto de buena calidad
  - Fuzzy matching con rapidfuzz para tolerar errores de OCR
  - Generador SSE para streaming fila a fila al frontend
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterator

import cv2
import numpy as np
import openpyxl
import pytesseract
import requests
from rapidfuzz import fuzz

# ── Configuración ──────────────────────────────────────────────────────────────

TESSERACT_CANDIDATES = [
    Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
    Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
]

EXCEL_DATE_ORIGIN = date(1899, 12, 30)

# Transformaciones de imagen ordenadas de mayor a menor probabilidad de éxito
# para cédulas colombianas (equivalentes al PHP con Intervention Image + GD)
TRANSFORM_NAMES: list[str] = [
    "binarized_base",     # escala de grises → OTSU  (base más confiable)
    "high_contrast",      # grises → CLAHE alto → OTSU
    "resize_2x",          # grises → OTSU → escalar ×2
    "grayscale_only",     # solo escala de grises (sin binarizar)
    "extra_threshold",    # threshold adaptativo gaussiano
    "fecha_optimizado",   # escalar ×3 → nitidez → OTSU  (optimizado para fechas)
    "sharpen",            # OTSU → unsharp mask
    "brightness",         # normalizar brillo → OTSU
    "fecha_erosion",      # escalar ×4 → OTSU → nitidez (erosión sintética)
]

# PSMs de Tesseract a probar por cada transformación
PSM_MODES: list[int] = [6, 3, 11, 4, 7, 8]

# Umbral de similitud (0–100) para fuzzy matching de palabras
FUZZY_WORD_THRESHOLD = 82

# Texto mínimo para parada anticipada (equivalente al break 2 del PHP)
EARLY_STOP_CHARS = 100
EARLY_STOP_WORDS = 10


# ── Tesseract ──────────────────────────────────────────────────────────────────

def configure_tesseract() -> None:
    for candidate in TESSERACT_CANDIDATES:
        if candidate.exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return
    raise FileNotFoundError("No se encontró tesseract.exe en las rutas esperadas.")


def _available_language() -> str:
    try:
        languages = set(pytesseract.get_languages(config=""))
    except pytesseract.TesseractError:
        return "eng"
    return "spa" if "spa" in languages else "eng"


# ── Preprocesamiento de imagen ─────────────────────────────────────────────────

def _otsu(gray: np.ndarray) -> np.ndarray:
    _, t = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return t


def _unsharp(img: np.ndarray, sigma: float = 1.0, strength: float = 1.5) -> np.ndarray:
    """Unsharp mask — aumenta nitidez conservando bordes."""
    blurred = cv2.GaussianBlur(img, (0, 0), sigma)
    return cv2.addWeighted(img, 1.0 + strength, blurred, -strength, 0)


def _safe_resize(gray: np.ndarray, factor: float, max_pixels: int = 4_000_000) -> np.ndarray:
    """Escala la imagen limitando el resultado a max_pixels para no saturar memoria."""
    h, w = gray.shape[:2]
    effective = min(factor, (max_pixels / (h * w)) ** 0.5)
    if effective <= 1.05:
        return gray
    return cv2.resize(gray, None, fx=effective, fy=effective,
                      interpolation=cv2.INTER_CUBIC)


def _deskew(gray: np.ndarray) -> np.ndarray:
    """
    Detecta y corrige la rotación del documento usando minAreaRect.
    Solo corrige ángulos entre 0.5° y 20° para evitar falsos positivos.
    """
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    coords = np.column_stack(np.where(binary > 0))
    if len(coords) < 50:
        return gray
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
    if abs(angle) < 0.5 or abs(angle) > 20:
        return gray
    h, w = gray.shape[:2]
    M = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1.0)
    return cv2.warpAffine(gray, M, (w, h),
                          flags=cv2.INTER_CUBIC,
                          borderMode=cv2.BORDER_REPLICATE)


def _apply_transform(img_bgr: np.ndarray, name: str) -> np.ndarray:
    """
    Aplica una de las 9 variantes de preprocesamiento sobre la imagen BGR.
    Equivalente a las funciones crearImagenTemporalIntervention del PHP.
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    if name == "grayscale_only":
        # PASO 1 del PHP: solo escala de grises
        return gray

    if name == "binarized_base":
        # PASO 1+2 del PHP: grises → binarización OTSU
        return _otsu(gray)

    if name == "high_contrast":
        # PASO 1+2+contraste_alto: CLAHE potente antes de OTSU
        clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
        return _otsu(clahe.apply(gray))

    if name == "resize_2x":
        # PASO 1+2+redimensionar: escalar ×2 (PHP resize w*2, h*2)
        return _otsu(_safe_resize(gray, 2.0))

    if name == "sharpen":
        # PASO 1+2+nitidez: unsharp mask sobre imagen binarizada
        return _unsharp(_otsu(gray))

    if name == "brightness":
        # PASO 1+2+brillo: normalizar rango dinámico → OTSU
        normalized = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
        return _otsu(normalized)

    if name == "extra_threshold":
        # PASO 1+2+umbral_extra: threshold adaptativo gaussiano
        return cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 15, 10
        )

    if name == "fecha_optimizado":
        # PHP fechas_optimizado: ×3 zoom + nitidez alta + contraste alto
        scaled = _safe_resize(gray, 3.0)
        sharpened = _unsharp(scaled, sigma=1.0, strength=2.0)
        return _otsu(sharpened)

    if name == "fecha_erosion":
        # PHP fechas_erosion: ×4 zoom + OTSU + nitidez fuerte
        scaled = _safe_resize(gray, 4.0)
        binarized = _otsu(scaled)
        return _unsharp(binarized, sigma=0.5, strength=1.5)

    return gray


def preprocess_and_ocr(image_path: Path, lang: str) -> str:
    """
    Aplica 9 transformaciones × 6 modos PSM de Tesseract.
    Devuelve el texto más largo encontrado.
    Parada anticipada (equivalente al break 2 del PHP) cuando se detecta
    texto de buena calidad: más de EARLY_STOP_CHARS caracteres y
    EARLY_STOP_WORDS palabras.
    """
    img_bgr = cv2.imread(str(image_path))
    if img_bgr is None:
        raise FileNotFoundError(f"No se pudo abrir la imagen: {image_path}")

    # Corrección de rotación antes de cualquier transformación
    base_gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    corrected = _deskew(base_gray)
    if corrected is not base_gray:
        img_bgr = cv2.cvtColor(corrected, cv2.COLOR_GRAY2BGR)

    best_text = ""

    for transform_name in TRANSFORM_NAMES:
        try:
            processed = _apply_transform(img_bgr, transform_name)
        except Exception:
            continue

        for psm in PSM_MODES:
            try:
                text = pytesseract.image_to_string(
                    processed, lang=lang,
                    config=f"--oem 3 --psm {psm}"
                ).strip()
            except Exception:
                continue

            if len(text) > len(best_text):
                best_text = text

            # Parada anticipada — texto de calidad suficiente encontrado
            if len(text) > EARLY_STOP_CHARS and len(text.split()) > EARLY_STOP_WORDS:
                return best_text

    return best_text


# ── Normalización de texto ─────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Minúsculas, sin tildes, sin caracteres especiales."""
    text = unicodedata.normalize("NFD", text.lower())
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9/ ]", " ", text)


# ── Fuzzy matching (rapidfuzz) ─────────────────────────────────────────────────

def _fuzzy_word_found(ocr_words: list[str], needle: str) -> bool:
    """
    Busca needle entre las palabras del OCR usando similitud de caracteres.
    Usa rapidfuzz.fuzz.ratio para comparar cada par.
    """
    for word in ocr_words:
        if fuzz.ratio(needle, word) >= FUZZY_WORD_THRESHOLD:
            return True
    return False


def _check_field(ocr_norm: str, value: Any) -> tuple[bool, str]:
    """
    Comprueba si el valor del campo aparece en el texto OCR normalizado.
    Primero intenta coincidencia exacta como subcadena, luego fuzzy por palabras.
    """
    if value is None:
        return False, ""
    needle = _normalize(str(value)).strip()
    if not needle:
        return False, str(value)

    # 1. Coincidencia exacta
    if needle in ocr_norm:
        return True, str(value)

    # 2. Fuzzy: todos los tokens del campo deben encontrarse en el OCR
    ocr_words = [w for w in ocr_norm.split() if w]
    needle_words = [w for w in needle.split() if w]
    if needle_words:
        found = all(_fuzzy_word_found(ocr_words, nw) for nw in needle_words)
        return found, str(value)

    return False, str(value)


# ── Fechas ─────────────────────────────────────────────────────────────────────

def _excel_serial_to_date(serial: int) -> date:
    return EXCEL_DATE_ORIGIN + timedelta(days=int(serial))


def _parse_excel_date(value: Any) -> tuple[list[str], str]:
    """
    Convierte el valor de la celda I (serial, str o datetime) en:
    - lista de variantes a buscar en el OCR
    - cadena display para mostrar en la tabla
    """
    d: date | None = None

    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            d = _excel_serial_to_date(int(value))
        except Exception:
            return [], str(value)
    elif isinstance(value, str):
        v = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                d = datetime.strptime(v, fmt).date()
                break
            except ValueError:
                continue

    if d is None:
        return [], str(value) if value else ""

    display = f"{d.day:02d}/{d.month:02d}/{d.year}"
    variants = [
        f"{d.day:02d}{d.month:02d}{d.year}",
        f"{d.day:02d}/{d.month:02d}/{d.year}",
        f"{d.day}/{d.month}/{d.year}",
        f"{d.day:02d}-{d.month:02d}-{d.year}",
        str(d.year),
    ]
    return variants, display


# ── Cálculo de coincidencias ───────────────────────────────────────────────────

def calculate_match(
    ocr_text: str,
    doc_num: Any,
    nombres: list[Any],
    birth: Any,
) -> dict:
    """
    Calcula porcentaje de coincidencia entre el OCR y los 6 campos del Excel.
    Usa exact + fuzzy matching para tolerar errores tipográficos del OCR.
    """
    ocr_norm = _normalize(ocr_text)
    results: dict[str, Any] = {}
    hits = 0
    total = 0

    # Número de documento (columna D)
    found, fmt = _check_field(ocr_norm, doc_num)
    results["documento"] = {"valor": fmt, "encontrado": found}
    total += 1
    hits += int(found)

    # Nombres y apellidos (columnas E, F, G, H)
    for label, val in zip(
        ["primer_nombre", "segundo_nombre", "primer_apellido", "segundo_apellido"],
        nombres,
    ):
        found, fmt = _check_field(ocr_norm, val)
        results[label] = {"valor": fmt, "encontrado": found}
        total += 1
        hits += int(found)

    # Fecha de nacimiento (columna I)
    date_variants, date_display = _parse_excel_date(birth)
    date_found = any(_normalize(v) in ocr_norm for v in date_variants)
    if not date_found and date_display:
        # Fallback fuzzy sobre la fecha completa
        ocr_words = [w for w in ocr_norm.split() if w]
        date_found = _fuzzy_word_found(ocr_words, _normalize(date_display))
    results["fecha_nacimiento"] = {"valor": date_display, "encontrado": date_found}
    total += 1
    hits += int(date_found)

    pct = round((hits / total) * 100) if total else 0
    results["porcentaje"] = pct
    results["estado"] = "OK" if pct >= 60 else "REVISAR"
    return results


# ── Procesamiento del Excel — generador SSE ────────────────────────────────────

def _get_cell(row: tuple, col: int) -> Any:
    idx = col - 1
    return row[idx] if idx < len(row) else None


def process_excel_stream(excel_path: Path, tmp_dir: Path) -> Iterator[dict]:
    """
    Generador que procesa el Excel fila a fila haciendo yield de cada resultado.
    Emite primero un evento 'meta' con el total de filas para la barra de progreso,
    luego un evento 'row' por cada fila procesada, y finalmente 'done'.
    """
    configure_tesseract()
    lang = _available_language()
    tmp_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    COL_D, COL_E, COL_F, COL_G, COL_H, COL_I = 4, 5, 6, 7, 8, 9
    COL_BB, COL_BC = 54, 55

    total_rows = max((ws.max_row or 1) - 1, 0)
    yield {"type": "meta", "total": total_rows}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        doc_num = _get_cell(row, COL_D)
        nombres = [
            _get_cell(row, COL_E),
            _get_cell(row, COL_F),
            _get_cell(row, COL_G),
            _get_cell(row, COL_H),
        ]
        birth = _get_cell(row, COL_I)
        url_bb = _get_cell(row, COL_BB)
        url_bc = _get_cell(row, COL_BC)

        # Saltar filas completamente vacías
        if not doc_num and not any(nombres):
            continue

        _, date_display = _parse_excel_date(birth)

        entry: dict[str, Any] = {
            "type": "row",
            "fila": row_idx,
            "documento": str(doc_num) if doc_num else "",
            "nombres": [str(n) if n else "" for n in nombres],
            "fecha_nacimiento": date_display or (str(birth) if birth else ""),
            "match": None,
            "error": None,
        }

        urls = [
            str(u) for u in [url_bb, url_bc]
            if u and str(u).startswith("http")
        ]

        if not urls:
            entry["error"] = "Sin URLs de imagen válidas"
            yield entry
            continue

        combined_text = ""
        img_errors: list[str] = []

        for i, url in enumerate(urls):
            img_path = tmp_dir / f"row{row_idx}_img{i}.jpg"
            try:
                resp = requests.get(
                    url,
                    headers={"User-Agent": "Mozilla/5.0"},
                    timeout=30,
                )
                resp.raise_for_status()
                img_path.write_bytes(resp.content)
                text = preprocess_and_ocr(img_path, lang)
                combined_text += " " + text
            except Exception as exc:
                img_errors.append(str(exc))
            finally:
                img_path.unlink(missing_ok=True)

        if not combined_text.strip():
            entry["error"] = "; ".join(img_errors) if img_errors else "No se extrajo texto"
            yield entry
            continue

        entry["match"] = calculate_match(combined_text, doc_num, nombres, birth)
        if img_errors:
            entry["error"] = f"Advertencia (imagen parcial): {img_errors[0]}"

        yield entry

    yield {"type": "done"}
