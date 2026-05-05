"""
Aplicación Flask — Verificador de documentos de identidad colombianos.
Usa Server-Sent Events (SSE) para enviar cada fila procesada en tiempo real.
"""
import io
import json
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, Response, render_template, request, send_file, stream_with_context

from processor import process_excel_stream

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
TMP_DIR = BASE_DIR / "uploads" / "tmp_images"
UPLOAD_DIR.mkdir(exist_ok=True)
TMP_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


def _allowed(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def _sse(data: dict) -> str:
    """Formatea un dict como evento SSE."""
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/procesar", methods=["POST"])
def procesar():
    if "archivo" not in request.files:
        return Response(
            _sse({"type": "error", "message": "No se recibió ningún archivo."}),
            mimetype="text/event-stream",
        ), 400

    file = request.files["archivo"]
    if not file.filename or not _allowed(file.filename):
        return Response(
            _sse({"type": "error", "message": "Solo se aceptan archivos .xlsx o .xlsm."}),
            mimetype="text/event-stream",
        ), 400

    safe_name = f"{uuid.uuid4().hex}{Path(file.filename).suffix.lower()}"
    excel_path = UPLOAD_DIR / safe_name
    file.save(str(excel_path))

    def generate():
        try:
            for event in process_excel_stream(excel_path, TMP_DIR):
                yield _sse(event)
        except Exception as exc:
            yield _sse({"type": "error", "message": str(exc)})
        finally:
            excel_path.unlink(missing_ok=True)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",   # evita buffering en nginx/proxies
        },
    )


@app.route("/descargar_excel", methods=["POST"])
def descargar_excel():
    data = request.get_json(force=True, silent=True) or {}
    rows = data.get("rows", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # ── Encabezados ──────────────────────────────────────────────────────────
    headers = [
        "Fila", "Documento",
        "Primer Nombre", "Segundo Nombre", "Primer Apellido", "Segundo Apellido",
        "F. Nacimiento",
        "Doc ✓", "N1 ✓", "N2 ✓", "A1 ✓", "A2 ✓", "Fec ✓",
        "Coincidencia %", "Estado", "Error",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Filas de datos ───────────────────────────────────────────────────────
    fill_ok  = PatternFill("solid", fgColor="D1E7DD")
    fill_rev = PatternFill("solid", fgColor="F8D7DA")
    fill_sin = PatternFill("solid", fgColor="E2E3E5")

    def _yesno(match_dict, field):
        """Devuelve 'Sí' / 'No' según encontrado; '—' si no hay match."""
        if not match_dict:
            return "—"
        info = match_dict.get(field)
        if info is None:
            return "—"
        return "Sí" if info.get("encontrado") else "No"

    for r in rows:
        m      = r.get("match") or {}
        nombres = r.get("nombres") or ["", "", "", ""]

        row_data = [
            r.get("fila", ""),
            r.get("documento", ""),
            nombres[0] if len(nombres) > 0 else "",
            nombres[1] if len(nombres) > 1 else "",
            nombres[2] if len(nombres) > 2 else "",
            nombres[3] if len(nombres) > 3 else "",
            r.get("fecha_nacimiento", ""),
            _yesno(m, "documento"),
            _yesno(m, "primer_nombre"),
            _yesno(m, "segundo_nombre"),
            _yesno(m, "primer_apellido"),
            _yesno(m, "segundo_apellido"),
            _yesno(m, "fecha_nacimiento"),
            m.get("porcentaje", "") if m else "",
            m.get("estado", "") if m else "",
            r.get("error", "") or "",
        ]
        ws.append(row_data)

        # Color de fila según estado
        estado = m.get("estado") if m else None
        if estado == "OK":
            row_fill = fill_ok
        elif estado == "REVISAR":
            row_fill = fill_rev
        else:
            row_fill = fill_sin

        for cell in ws[ws.max_row]:
            cell.fill = row_fill

    # ── Ajustar ancho de columnas ────────────────────────────────────────────
    col_widths = [6, 14, 16, 16, 16, 16, 14, 7, 7, 7, 7, 7, 7, 14, 10, 40]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ── Serializar y enviar ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"verificacion_{timestamp}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    # threaded=True es necesario para que el streaming funcione correctamente
    app.run(debug=True, threaded=True)
